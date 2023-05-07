# _*_coding:utf-8_*_
import os
from os import walk
from os.path import join, dirname

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm


# @PROJECT : DY_SCRM_TEST_PROJECT
# @Time : 2023/3/3 17:00
# @Author : Byseven
# @File : 标注表格中两个sheet非重复值.py
# @SoftWare:


def contrast_sheets(excel_path: str, ts_sheet_name: str, ts_column: str, ts_max: int, p_sheet_name: str, p_column: str,
                    pin_max: int):
    """
    用于过滤两个sheet中对应两列数据的非重复值
    ！！请确保两列数据格式的一致性！！，两列数据的唯一性，选中一列，标记重复值，去掉
    对比两个sheet中的B列【原本用于匹配千川视频id】，将两者不包含的分别标注
    :param excel_path: 表格路径
    :param ts_sheet_name: tssheet名称
    :param ts_column: ts列
    :param ts_max: ts最大行数
    :param p_sheet_name: 品牌sheet名称
    :param p_column: 品牌列
    :param pin_max: 品牌最大行数
    :return:
    """
    fill = PatternFill("solid", fgColor="1874CD")  # 蓝色样式标记-品牌
    fill2 = PatternFill("solid", fgColor="FF0000")  # 红色样式标记-TS
    wb = load_workbook(excel_path)
    print('🚀读取表格成功！请中途勿关闭程序！\n--会导致表格损坏！！')
    ws = wb[ts_sheet_name]
    print(f'🚀读取sheet:{ts_sheet_name}')
    t_data = []  # 存放sheet1数据
    for i in tqdm(range(2, ts_max + 1)):
        v_id = ws[f'{ts_column}{i}'].value
        # print(f'读取TS第{i}行 - [{v_id}]')
        t_data.append(v_id)
    print(f'🚀读取{ts_sheet_name}完成！\n')
    wp = wb[p_sheet_name]
    print(f'🚀读取sheet:{p_sheet_name}')
    p_data = []  # 存放sheet2数据
    for j in tqdm(range(2, pin_max + 1)):
        vp_id = wp[f'{p_column}{j}'].value
        # print(f'读取PIN第{j}行 - [{vp_id}]')
        p_data.append(vp_id)
    print(f'🚀读取{p_sheet_name}完成！\n')

    print(f'🚀正在标注  {ts_sheet_name}  与  {p_sheet_name}  差异！')
    for tid in tqdm(t_data):
        if tid not in p_data:
            sty = t_data.index(tid) + 2
            ws[f'{ts_column}{sty}'].fill = fill2
    print(f'🚀标注{ts_sheet_name}完成！\n')

    print(f'🚀正在标注  {p_sheet_name}  与  {ts_sheet_name}  差异！')
    for pid in tqdm(p_data):  # 遍历品牌数据
        if pid not in t_data:  # 如果值不在TS列表中
            ty = p_data.index(pid) + 2
            wp[f'{p_column}{ty}'].fill = fill
    print(f'🚀标注 {p_sheet_name} 完成！\n')
    try:
        wb.save(excel_path)
        print('🚀处理完成')
    except IOError as e:
        print(f'！！保存出错！！：{e},\n 请 确 认 没 有 其 他 应 用 打 开 此 表 格')


def get_file_list(path: str) -> list:
    """
    遍历返回子文件
    :param path:
    :return:
    """
    file_list = []
    for home, dirs, files in walk(path):
        for filename in files:
            # 文件名列表，包含完整路径
            file_list.append(join(home, filename))
            # # 文件名列表，只包含文件名
            # Filelist.append( filename)
    return file_list


def t():
    current_work_dir = dirname(__file__)  # 当前文件所在的目录
    return current_work_dir
    # weight_path = os.path.join(current_work_dir, weight_path)  # 再加上它的相对路径，这样可以动态生成绝对路径


def get_file_name(path):
    file_name = path.split('\\')
    return file_name.pop()


def clear():
    os.system('cls')


def do_it():
    print('\n\n┌---🚀本工具默认从第二行开始对比！---┐')
    print('|                                    |')
    l = get_file_list(t())
    for i in l:
        check = i[:-5:-1][::-1]
        if check == 'xlsx':
            ch = 1
            e = ''
            try:
                ch = int(input(
                    f'|{" " * 10}✅1.检索表格{" " * 14}|\n|{" " * 10}❎2.退出程序{" " * 14}|\n|{" " * 36}|\n└{"-" * 36}┘\n📝请输入:'))
            except ValueError as e:
                pass
            if ch == 1:
                excel_file = get_file_name(i)
                print(f'🚀读取到表格： {excel_file} ')
                try:
                    ts_name, ts_col, ts_ma = input('📝请输入sheet1名称(空格)列名(空格)最大行数：').split()
                    try:
                        p_name, p_col, p_ma = input('📝请输入sheet2名称(空格)列名(空格)最大行数：').split()
                        contrast_sheets(excel_path=i,
                                        ts_sheet_name=ts_name, ts_column=ts_col, ts_max=int(ts_ma),
                                        p_sheet_name=p_name, p_column=p_col, pin_max=int(p_ma)
                                        )
                    except ValueError as e:
                        clear()
                        print(f'{e}\n🔴请输入正确的表格信息')
                        return do_it()
                except ValueError as e:
                    clear()
                    print(f'{e}\n🔴请输入正确的表格信息')
                    return do_it()

            elif ch == 2:
                print('🟢感谢使用')
                exit()
            else:
                clear()
                print(f'{e}\n🔴请输入指定数字选择')
                return do_it()


if __name__ == '__main__':
    do_it()
