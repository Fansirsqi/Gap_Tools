from openpyxl import load_workbook
from collections import Counter
from utils import *
def set_title(path):
    wb = load_workbook(path)
    ws_sys = wb['system']
    ws_bra = wb['brand']
    _s_title = [cell.value for cell in ws_sys[1]]
    _b_title = [cell.value for cell in ws_bra[2]]
    a = Counter(_s_title)
    sys_line1 = ws_sys[1]
    
    # for t in _s_title:
    #     if t in _b_title:
    #         from_sys = ws_sys.cell(row=1,column=_s_title[t])
    #         print()


path = file_selector()
set_title(path)
