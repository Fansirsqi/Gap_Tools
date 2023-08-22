from openpyxl.styles import PatternFill,Font

# 创建条件格式规则
brand_fill = PatternFill(start_color="eeac6c", end_color="eeac6c", fill_type="solid")
system_fill = PatternFill(start_color="12a107", end_color="12a107", fill_type="solid")#绿色
gap_fill = PatternFill(start_color="ff8787", end_color="ff8787", fill_type="solid")

nothing_fill = PatternFill(start_color="AEAEAE",end_color="AEAEAE",fill_type="solid")# 无需GAP字段填充
gap_fill_less = PatternFill(start_color="6e8bea", end_color="6e8bea", fill_type="solid")#紫色->小于-0.005
gap_fill_greater = PatternFill(start_color="E64848", end_color="E64848", fill_type="solid")#红色->大于0.005
gap_fill_false = PatternFill(start_color="f44444", end_color="f44444", fill_type="solid")#红色

gap_font = Font(
    name="Calibri",  # 字体
    size=8,         # 字体大小
    color="000000",  # 字体颜色，用16进制rgb表示
    bold=False,       # 是否加粗，True/False
    italic=False,     # 是否斜体，True/False
    strike=None,     # 是否使用删除线，True/False
    underline=None,  # 下划线, 可选'singleAccounting', 'double', 'single', 'doubleAccounting'
)
title_font = Font(
    name="等线",  # 字体
    size=10,         # 字体大小
    color="000000",  # 字体颜色，用16进制rgb表示
    bold=True,       # 是否加粗，True/False
    italic=False,     # 是否斜体，True/False
    strike=None,     # 是否使用删除线，True/False
    underline=None,  # 下划线, 可选'singleAccounting', 'double', 'single', 'doubleAccounting'
)