# _*_coding:utf-8_*_
from logs import logger
import re
import time
import os
from functions import auto_save, get_data_by_any_row, set_gap
from utils import input_selector, file_selector, clear
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rich.progress import track


class Gap:
    """包含wb[openpyxl],pw[pandas]"""

    def __init__(self, excel_path: str):
        """
        传入表格路径
        @param excel_path:
        """
        logger.info("开始读取表格,请耐心等待")
        if os.path.exists(excel_path):
            self.excel_path = excel_path
            self.wb = load_workbook(self.excel_path)
            self.pwb = pd.ExcelFile(self.excel_path)
            logger.success(f"已成功读取表格: {excel_path}")
        else:
            logger.error(f"{excel_path}路径文件不存在！")


@auto_save
def get_data_by_row_title(gap: Gap):
    logger.warning(
        """
在这之前！你需要将品牌表头第2行,改成TS对应的mapping值！！这很重要
Mapping from sheet title data
匹配两个表头,将能匹配上的,数据传入另一个表头下方
@param ws_name:品牌sheetName
@param _is_reference: 1/0 (是/否开启全部匹配)
传参方式 品牌表名 1/0
    """
    )
    _names = input_selector("请分别输入【品牌表名  ？】")
    barnd, system = _names[:2]
    get_data_by_any_row(brand_sheet=gap.wb[barnd], system_sheet=gap.wb[system])


@auto_save
def contrast_sheets(gap: Gap):
    """标注两列数据的差异
    Args:
        sheet_name1 (str): sheet1名称
        ts_column (str): sheet1_列名
        sheet_name2 (str): sheet2名称
        p_column (str): sheet2列名
    """
    logger.info(
        "注意!!操作的数据列请确保格式完全一致,请排除空格引号等问题！\n" * 3,
        color="random",
    )
    _val1 = input_selector("请分别输入【表名1 匹配列名1(需要大写字母)】")
    try:
        sheet_name1, column_1 = _val1[:2]
        logger.info("载入表1,成功")
    except Exception as e:
        logger.info("载入表1错误", e)
        exit(1)
    _val2 = input_selector("请分别输入【表名2 匹配列名2(需要大写字母)】")
    try:
        sheet_name2, column_2 = _val2[:2]
        logger.info("载入表2,成功")
    except Exception as e:
        logger.info("载入表2错误", e)
        exit(1)
    fill = PatternFill("solid", fgColor="1874CD")  # 蓝色样式标记-品牌
    fill2 = PatternFill("solid", fgColor="FF0000")  # 红色样式标记-TS
    ws = gap.wb[sheet_name1]
    wp = gap.wb[sheet_name2]
    t_data = set(ws[column_1][2 : ws.max_row])
    _tdata = [cell.value for cell in t_data]
    p_data = set(wp[column_2][2 : wp.max_row])
    _pdata = [cell.value for cell in p_data]
    for sty, tid in track(enumerate(t_data, start=2)):
        # logger.info(sty,tid.value)
        if tid.value not in _pdata:
            ws[f"{column_1}{sty}"].fill = fill2
    for py, pid in track(enumerate(p_data, start=2)):
        if pid.value not in _tdata:
            wp[f"{column_2}{py}"].fill = fill


@auto_save
def set_gaps(gap: Gap):
    """_生成GAP表,无敌_

    Args:
        gap (Gap): _description_
    """
    _names = input_selector("请输入要GAP的表名:品牌 TS 标题行")
    pn, tn, tl = _names[:3]
    pw = gap.wb[pn]
    tw = gap.wb[tn]
    set_gap(pws=pw, tws=tw, title_row=int(tl))


get_data_by_row_title.__name__ = "快速引用品牌列数据"
contrast_sheets.__name__ = "对比两列数据并标记"
set_gaps.__name__ = "生成GAP表[电商生意low out]"


fundict = {"1": get_data_by_row_title, "2": contrast_sheets, "3": set_gaps}


def function_list(obj: Gap):
    print(f'{"="*10}func list{"="*10}')
    for i in fundict:
        print(f" {i}:{fundict[i].__name__}")
    print(f'{"="*10}func list{"="*10}')
    _x = input_selector("选择功能：")
    x = _x[:1][0]
    if x not in fundict.keys():
        clear()
        print("输入有误！,请重新选择")
        return function_list()
    else:
        # try:
        return fundict[x](obj)
        # except Exception as e:
        #     Log.error(e)
        #     exit(1)


if __name__ == "__main__":
    path = file_selector()
    logger.info(path)
    obj = Gap(path)
    function_list(obj)
