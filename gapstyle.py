from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Alignment, Font
from pydantic import BaseModel
from typing import Dict, Any, ClassVar, List, Optional, Union


class FillConfig(BaseModel):
    fill_1: ClassVar[PatternFill] = PatternFill(end_color="00EE822F", fill_type="solid")
    """大于10%-填充色 cell.fill"""
    fill_2: ClassVar[PatternFill] = PatternFill(end_color="00FFEFC1", fill_type="solid")
    """介于0.5%到10%-填充色 cell.fill"""
    fill_3: ClassVar[PatternFill] = PatternFill(end_color="00FCDECD", fill_type="solid")
    """介于-0.5%到-10%-填充色 cell.fill"""
    fill_4: ClassVar[PatternFill] = PatternFill(end_color="00F4B382", fill_type="solid")
    """小于-10%-填充色 cell.fill"""
    brand_title_fill: ClassVar[PatternFill] = PatternFill(start_color="c55c10", end_color="c55c10", fill_type="solid")
    """品牌表头填充色"""
    system_title_fill: ClassVar[PatternFill] = PatternFill(start_color="248e87", end_color="248e87", fill_type="solid")
    """系统表头填充色"""
    gap_title_fill: ClassVar[PatternFill] = PatternFill(start_color="c91d32", end_color="c91d32", fill_type="solid")
    """gap表头填充色"""
    all_title_fill: ClassVar[PatternFill] = PatternFill(start_color="91aadf", end_color="91aadf", fill_type="solid")
    """通用表头填充色"""


class FontConfig(BaseModel):
    all_title_font: ClassVar[Font] = Font(color="ffffff", bold=True, name="微软雅黑", size=12)
    """通用表头字体"""
    gap_title_font: ClassVar[Font] = Font(name="微软雅黑", bold=True, color="FFFFFF", size=9)
    """gap表头字体"""
    gap_filed_font: ClassVar[Font] = Font(size=8, color="000000", name="微软雅黑")
    """gap表 字段-字段字体"""


font_config = FontConfig()
fill_config = FillConfig()
rules = []
"""规则列表"""

# 创建DataBarRule规则
rule1 = CellIsRule(operator=">", formula=[0.1], fill=fill_config.fill_1)
"""大于10%"""
rules.append(rule1)
rule2 = CellIsRule(operator="between", formula=[0.005, 0.1], fill=fill_config.fill_2)
"""介于0.5%到10%"""
rules.append(rule2)
rule3 = CellIsRule(operator="between", formula=[-0.1, -0.005], fill=fill_config.fill_3)
"""介于-0.5%到-10%"""
rules.append(rule3)
rule4 = CellIsRule(operator="<", formula=[-0.005], fill=fill_config.fill_4)
"""小于-10%"""
rules.append(rule4)

center_align = Alignment(horizontal="center", vertical="center")
"""居中对其 使用cell.alignment="""

if __name__ == "__main__":
    pass
    # for rule in rules:
    # ws.conditional_formatting.add("C1:C38", rule)
