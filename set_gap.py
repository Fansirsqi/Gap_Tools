from openpyxl.styles import numbers
from openpyxl import load_workbook
from utils import *
from Sheetstyle import *

from log import do_logger

def set_gap_vlookup(path):
    wb = load_workbook(path)
    ws_brand = wb['brand_data']
    
    ws_system = wb['system']
    ws_gap = wb.create_sheet('gap')
    #取最大列，最大行
    max_col = max(ws_brand.max_column, ws_system.max_column)
    max_row = max(ws_brand.max_row, ws_system.max_row)
    min_row = min(ws_brand.max_row, ws_system.max_row)
    sr = set_az()
    system_title:dict = {cell.value:_i for _i,cell in enumerate(ws_system[1],start=1)}#此处将字段作为键，索引作为值
    _is_all:bool = False #匹配行数True = all,False = 4
    for i in tqdm(range(1,max_col*3,3),ncols=120,desc='gap进度'):#此处i代表新表中的每一列的开头一列
        # print(i)
        brand_col = i
        system_col = i+1
        gap_col = i+2
        _index = i//3
        quote_col_name = sr[_index] #需要引入的列名
        # 当前列title,这里需要直接取brand字段
        title = ws_brand[f'{quote_col_name}2'].value
        for irow in range(1,min_row+1):
            _brand = ws_brand[f'{quote_col_name}{irow}']
            _system = ws_system[f'{quote_col_name}{irow}']
            # 标注出3个cell对象
            gap_brand_cell = ws_gap.cell(row=irow,column=brand_col)
            gap_system_cell = ws_gap.cell(row=irow,column=system_col)
            gap_gap_cell = ws_gap.cell(row=irow,column=gap_col)
            #注意不用区拿标题，所以当irow=1时，需要过滤
            F1 = gap_system_cell.coordinate
            G1 = gap_brand_cell.coordinate
            H1 = gap_gap_cell.coordinate
            # letter_part = H1.rstrip('0123456789')
            # use_letter = f'{letter_part}:{letter_part}'
            # 前两列需要引入的公式
            _col_name = f'{quote_col_name}{irow}'
            if irow == 1:#如果是首行
                quite_brand = f'={ws_brand.title}!{_col_name}&"-品牌"'#品牌表头
                quite_system = f'={ws_system.title}!{_col_name}&"-系统"'#系统表头
                quite_gap = f'={ws_system.title}!{_col_name}&"-GAP"'
                gap_brand_cell.fill = brand_fill
                gap_system_cell.fill = system_fill
                gap_gap_cell.fill = gap_fill
                gap_brand_cell.font = title_font
                gap_system_cell.font = title_font
                gap_gap_cell.font = title_font
            else:#非首行
                if irow == 2:
                    quite_gap = title
                    gap_gap_cell.value = quite_gap
                    gap_gap_cell.fill = nothing_fill#标注无需gap字段
                    gap_brand_cell.fill = nothing_fill
                    gap_system_cell.fill = nothing_fill
                    continue
                if _is_all:
                    do_lines = max_row
                else:
                    do_lines = 4
                if irow < do_lines:
                        if title in system_title:#这里默认会检索key
                            _vindex = system_title[title]
                        else:
                            _vindex = 'no'
                        if _vindex == 'no':
                            quite_gap = None #Gap--------------------------
                            quite_system = None #系统
                            quite_brand = None # Brand
                            gap_gap_cell.fill = nothing_fill#标注无需gap字段
                            gap_brand_cell.fill = nothing_fill
                            gap_system_cell.fill = nothing_fill
                        else:
                            gap_brand_cell.font = gap_font
                            gap_system_cell.font = gap_font
                            gap_gap_cell.font = gap_font
                            do_logger(_brand.value,type(_brand.value),G1)
                            if isinstance(_brand.value,(int,float)):
                                # =IF(A1=0,IF(B1=0,0,IF(A1>B1,(B1-A1)/A1,(B1-A1)/B1)),IF(B1=0,0,IF(A1>B1,(B1-A1)/A1,(B1-A1)/B1)))
                                quite_gap = f'=IF({F1}=0,IF({G1}=0,0,IF({F1}>{G1},({G1}-{F1})/{F1},({G1}-{F1})/{G1})),IF({G1}=0,0,IF({F1}>{G1},({G1}-{F1})/{F1},({G1}-{F1})/{G1})))'#gap
                                gap_gap_cell.number_format = numbers.FORMAT_PERCENTAGE_00#设置百分比格式？
                            else:
                                quite_gap = f'=EXACT({F1},{G1})'
                            quite_system = f'=VLOOKUP(A{irow},{ws_system.title}!A:ZZ,{_vindex},0)'#系统
                            quite_brand = f'={ws_brand.title}!{_col_name}'#品牌---------------------
                else:#大于4行不处理
                    continue
            gap_brand_cell.value = quite_brand
            gap_system_cell.value = quite_system
            gap_gap_cell.value = quite_gap
    # wrokgap = wb.copy_worksheet(ws_gap)
    # wrokgap.title = 'wrokgap'
    print('🎁Gap_Sheet生成成功🎁，等待保存')
    wb.save(path)
    print('🎁OK')

def copy_sheet(path):
    wb = load_workbook(path,data_only=True)
    ws = wb['brand1']
    ws_data = wb.copy_worksheet(ws)
    ws_data.title = 'brand_data'
    wb.save(path)
    print('🎁OK')

path = file_selector()
# copy_sheet(path)
set_gap_vlookup(path)