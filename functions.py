from logs import logger
from collections import defaultdict
import re
from rich.progress import track
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from gapstyle import rules, center_align, font_config, fill_config


def auto_save(func):
    def saver(*args, **kwargs):
        data = func(*args, **kwargs)
        try:
            logger.info("开始数据保存,请耐心等待。。")
            args[0].wb.save(args[0].excel_path)
            logger.success("保存成功 🟢")  # noqa: F541)
        except IOError as e:
            logger.error(f"🔴保存失败🔴==>{e}")
        return data

    return saver


def get_data_by_any_row(system_sheet: Worksheet, brand_sheet: Worksheet, _is_reference: int = 0):
    """从表头,拿表头"""

    _is_reference = int(_is_reference)
    if _is_reference:
        logger.debug("开启模糊匹配！")
    else:
        logger.debug("未开启")
    logger.info(f"🟡开始匹配表头🟡{system_sheet.title}->{brand_sheet.title}1")

    brand_max_row = brand_sheet.max_row

    # 创建新表
    wb = brand_sheet.parent
    ts_sheet = wb.create_sheet(f"{brand_sheet.title}_new")

    # 获取品牌表第二行数据
    brand_title = [cell.value for cell in brand_sheet[2]]
    # 获取系统表头字段
    system_title = [cell.value for cell in system_sheet[1]]
    # 统计每个系统表头字段的出现次数
    counter = defaultdict(int)
    for cell in system_sheet[1]:
        counter[cell.value] += 1
    # 遍历system_title,并为重复的字段添加后缀
    unique_system_title = []
    for title in system_title:
        if counter[title] > 1:
            suffix = str(counter[title])
            counter[title] -= 1
            title += suffix
        unique_system_title.append(title)
    for _title in track(unique_system_title, description="匹配进度"):  # 遍历系统表头
        index_system = unique_system_title.index(_title)  # Bxx,Axx的数据 eg:A1
        ts_cell = ts_sheet.cell(row=1, column=index_system + 1)
        # --》第一行引入TS表头
        ts_cell.value = f"={system_sheet.title}!{ts_cell.coordinate}"
        v = re.sub(r"\d", "", _title)
        ts_sheet.cell(row=2, column=index_system + 1).value = v  # 第二行,表头,文本内容
        if _title in brand_title:  # 如果系统表头在品牌第二行,这里默认只匹配第一个字段,如果有相同字段
            index_brand = brand_title.index(_title)  # 索引-品牌
            _row = 3  # 第三行
            brand_cell = brand_sheet.cell(row=_row, column=index_brand + 1)  # 品牌第3行数据cell对象,kais
            logger.info(brand_cell.coordinate, brand_max_row)
            ts_cell_2 = ts_sheet.cell(row=_row, column=index_system + 1)
            ts_cell_2.value = f"={brand_sheet.title}!{brand_cell.coordinate}"
            logger.debug(f"匹配到 {_title} 在 {system_sheet.title}表头中")
        else:  # 如果系统表头NOT 在品牌第二行
            ts_cell_2 = ts_sheet.cell(row=2, column=index_system + 1)
            ts_cell_2.value = "/"
    logger.info("✨匹配完成✨")


def get_value_of_line(gws: Worksheet, line: int) -> list:
    """获取指定行的数据

    Args:
        gws (Worksheet): _description_
        line (int): _description_

    Returns:
        _type_: _description_
    """
    row_values = []
    for cell in gws[line]:
        row_values.append(cell.value)
    return row_values


def filed_comment(wb: Workbook, row_values: list, set_row: int = 1):
    """将字段来源也一并拿过去
    row_values:备注行数据
    set_row: 设置读取的行号,并写入目标行号
    """
    gws = wb["GAP"]
    for clo in range(1, len(row_values) * 3, 3):
        key = clo // 3
        one = clo
        three = clo + 2
        f_value = row_values[key]
        g_title = gws.cell(row=set_row, column=one)
        g_title.value = f_value
        gws.merge_cells(title_row=set_row, start_column=one, end_row=set_row, end_column=three)


def set_gap(pws: Worksheet, tws: Worksheet, title_row: int = 1):
    """生成GAP表头部

    Args:
        row_values (list): 传入TS/品牌表头数据列表
        pws (Worksheet): 品牌表对象
        tws (Worksheet): TS表对象
        title_row (int, optional): _表头所在行_. 默认值 2.
    """

    row_values = get_value_of_line(pws, title_row)  # 获取表头数据,从品牌表拿
    gws: Worksheet = pws.parent.create_sheet("GAP")
    # gws = wb["NewSheet"]

    for clo in range(1, len(row_values) * 3, 3):
        key = clo // 3
        one = clo
        two = clo + 1
        three = clo + 2
        f_value = row_values[key]
        logger.debug(f"载入[传入]字段 -- {f_value}")
        g_title = gws.cell(row=title_row, column=one)
        g_title.value = f_value
        g_title.fill = fill_config.all_title_fill
        g_title.font = font_config.all_title_font
        g_title.alignment = center_align
        try:  # 合并单元格
            gws.merge_cells(start_row=title_row, start_column=one, end_row=title_row, end_column=three)  # 这里的字段是来自row_list
            logger.debug(f"合并单元格成功 -- {f_value}")
        except Exception as e:
            logger.error(f"合并单元格失败 -- {e}")
        g_p = gws.cell(row=title_row + 1, column=one)
        g_t = gws.cell(row=title_row + 1, column=two)
        g_g = gws.cell(row=title_row + 1, column=three)

        _p_quote = pws.cell(row=title_row, column=key + 1).coordinate  # 此处引用的是品牌sheet,表头那一行
        _t_quote = tws.cell(row=title_row, column=key + 1).coordinate

        p_quote = f'={pws.title}!{_p_quote}&"-{pws.title}"'
        t_quote = f'={tws.title}!{_t_quote}&"-{tws.title}"'

        g_p.value = p_quote
        g_t.value = t_quote
        g_g.value = "Gap"

        g_p.fill = fill_config.brand_title_fill
        g_t.fill = fill_config.system_title_fill
        g_g.fill = fill_config.gap_title_fill

        g_p.font = font_config.gap_title_font
        g_t.font = font_config.gap_title_font
        g_g.font = font_config.gap_title_font

        g_p.alignment = center_align
        g_t.alignment = center_align
        g_g.alignment = center_align

    set_gap_title_value(pws, tws, gws, title_row + 2)


def set_gap_title_value(pws: Worksheet, tws: Worksheet, gws: Worksheet, set_row: int):
    """_设置GAP标题下的一行,附带格式_

    Args:
        pws (Worksheet): _description_
        tws (Worksheet): _description_
        gws (Worksheet): _description_
        set_row (int, optional): _description_..
    """
    for clo in range(1, pws.max_column * 3, 3):
        key = clo // 3
        one = clo
        two = clo + 1
        three = clo + 2

        g_p = gws.cell(row=set_row, column=one)
        g_t = gws.cell(row=set_row, column=two)
        g_g = gws.cell(row=set_row, column=three)
        # =IF(C2=0,IF(D2=0,0%,-100%),IF(D2=0,100%,(C2-D2)/D2))
        _g_t = g_t.coordinate
        _g_p = g_p.coordinate
        _g_g = g_g.coordinate

        _p_quote = pws.cell(row=set_row - 1, column=key + 1).coordinate
        _t_quote = tws.cell(row=set_row - 1, column=key + 1).coordinate

        p_quote = f"='{pws.title}'!{_p_quote}"
        t_quote = f"='{tws.title}'!{_t_quote}"
        gap_quote = f"=IF({_g_t}=0,IF({_g_p}=0,0%,-100%),IF({_g_p}=0,100%,({_g_p}-{_g_t})/{_g_t}))"
        """
        Gap公式:
        (TS-品牌)/品牌
        """

        g_p.value = p_quote
        g_t.value = t_quote
        g_g.value = gap_quote

        g_p.font = font_config.gap_filed_font
        g_t.font = font_config.gap_filed_font
        g_g.font = font_config.gap_filed_font

        g_p.alignment = center_align
        g_t.alignment = center_align
        g_g.alignment = center_align

        g_g.number_format = "0.00%"
        if one == 1:
            g_p.number_format = "YYYY/M/D"
            g_t.number_format = "YYYY/M/D"
        for rule in rules:
            gws.conditional_formatting.add(_g_g, rule)
