from openpyxl import load_workbook
from utils import *
from Sheetstyle import *
import time

path = './TS.xlsx'
wb = load_workbook(path,data_only=True) 
_ws_gap = wb['wrokgap']
_gap_title = {cell.value:i for i,cell in enumerate(_ws_gap[2],start=1)}
print(_gap_title)
max_rows = _ws_gap.max_row
for icol in _gap_title:
    if icol is None or icol == '/':
        continue
    else:
        for irow in range(3,max_rows):
            c = _gap_title[icol]
            gp_cell = _ws_gap.cell(row=irow,column=c)
            # print(gp_cell.value)
            if gp_cell.value is not None:
                # print(gp_cell.value,type(gp_cell.value))
                if gp_cell.value == '#N/A':
                    gp_cell.fill = gap_fill_false
                if isinstance(gp_cell.value,(int,float)) and not isinstance(gp_cell.value,(bool)):
                    if gp_cell.value < -0.005:#小于使用黄色
                        gp_cell.fill = gap_fill_less
                    elif gp_cell.value > 0.005:#大于使用红色
                        gp_cell.fill = gap_fill_greater
                else:
                    if not gp_cell.value:
                        gp_cell.fill = gap_fill_false
                        # print(gp_cell.value)
                        # time.sleep(1)
                    
                    

wb.save(path)
print('set style ok')